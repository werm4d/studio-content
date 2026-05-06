"""
M4D Chase CSV Importer
Merge 4 Design LLC — bookkeeping automation tool

Reads a Chase bank CSV export (downloaded from chase.com) and maps transactions
into the M4D annual bookkeeping workbook (01_Transactions tab).

Usage
-----
  # Basic — import all years found in the CSV:
  python3 M4D_Chase_Importer.py Chase6378_Activity_20260130.CSV --workbook Merge4Design_2026_Bookkeeping_v1.xlsx

  # Filter to a single tax year:
  python3 M4D_Chase_Importer.py Chase6378_Activity_20260130.CSV --workbook Merge4Design_2026_Bookkeeping_v1.xlsx --year 2026

  # Specify a custom output path (default: adds _imported suffix to workbook name):
  python3 M4D_Chase_Importer.py Chase6378_Activity_20260130.CSV --workbook path/to/workbook.xlsx --output path/to/output.xlsx

Rules engine order
-----------------
  1. Type-based rules  (ACCT_XFER → Transfer: Between Accounts, etc.)
  2. Keyword rules     (FEDEX → Printing & Repro, INSUREON → Insurance, etc.)
  3. Payee extraction  (BILLPAY "To <Name>" → extracts payee for Members/1099 matching)
  4. Member detection  (BILLPAY/MISC_CREDIT to partner name → Equity: Owner Distribution/Contribution)
  5. FLAGGED           (anything unmatched → orange highlight, needs manual review)

Output
------
  Writes matched rows into 01_Transactions tab of a copy of the workbook.
  - Light green cell  = auto-categorized (verify but likely correct)
  - Orange cell       = flagged, needs your decision before closing the books
  - Duplicate rows    = skipped automatically (checks date + amount against existing data)
  - Transfers         = imported but excluded from P&L by category

Repo: github.com/werm4d/merge4design
"""

import argparse
import csv
import os
import re
import sys
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── CONSTANTS ──────────────────────────────────────────────────────────────────
ACCOUNT_NAME = "Chase ...6378"

# Known M4D member names — exact match against extracted payee
MEMBERS = ["David Ainsworth", "Trevor Knight", "Leo Lynch", "Travis Davis"]

# ── CATEGORY MAPPING ENGINE ────────────────────────────────────────────────────
# Each rule: (priority, match_fn, category, payee_hint, is1099_hint)
# match_fn receives (row_dict, desc_upper, type_str, amount_float)
# Returns category string or None

def make_rules():
    return [

        # ── 1. TRANSFERS (always skip P&L) ──────────────────────────────────────
        (10, lambda d,u,t,a: t in ("ACCT_XFER",) or
             "ONLINE TRANSFER TO MMA" in u or
             "ONLINE TRANSFER FROM MMA" in u or
             "ONLINE TRANSFER FROM CHK" in u or
             "ONLINE TRANSFER TO CHK" in u,
         "Transfer: Between Accounts", None, "No"),

        # ── 2. WIRE FEES / BANK FEES ────────────────────────────────────────────
        (20, lambda d,u,t,a: t == "FEE_TRANSACTION" or
             "WIRE FEE" in u or "OFFICIAL CHECKS CHARGE" in u or
             "SERVICE FEE" in u or "MONTHLY FEE" in u or "NSF" in u,
         "Expense: Bank Fees", None, "No"),

        # ── 3. WIRE OUTGOING — check payee ──────────────────────────────────────
        # Member wire-outs are equity distributions; vendor wire-outs are expenses
        # Default here; member detection pass will override if payee matches member
        (15, lambda d,u,t,a: t == "WIRE_OUTGOING" and a < 0,
         "Equity: Member Loan Out", None, "No"),

        # ── 4. MISC_DEBIT (cash withdrawals) ────────────────────────────────────
        (16, lambda d,u,t,a: t == "MISC_DEBIT" or "WITHDRAWAL" in u,
         "Equity: Member Loan Out", None, "No"),

        # ── 5. INCOME — ACH credits from known clients ──────────────────────────
        (30, lambda d,u,t,a: "L.E.E.P." in u and t == "ACH_CREDIT",
         "Income: Design/Consulting Fees", "L.E.E.P.", "Yes"),

        # ── 6. INCOME — deposits / check deposits (positive amounts) ────────────
        (35, lambda d,u,t,a: t in ("DEPOSIT", "CHECK_DEPOSIT", "DSLIP") and a > 0
             and "INTEREST" not in u,
         "Income: Design/Consulting Fees", None, ""),

        # ── 7. INTEREST ─────────────────────────────────────────────────────────
        (35, lambda d,u,t,a: "INTEREST PAYMENT" in u or "INTEREST EARNED" in u,
         "Income: Other (Interest/Refunds)", None, "No"),

        # ── 8. CREDIT RETURNS (reversed payments back in) ───────────────────────
        (35, lambda d,u,t,a: "CREDIT RETURN" in u and a > 0,
         "Equity: Owner Contribution", None, "No"),

        # ── 9. MISC_CREDIT that isn't a return — default income ─────────────────
        (36, lambda d,u,t,a: t == "MISC_CREDIT" and a > 0,
         "Income: Other (Interest/Refunds)", None, "No"),

        # ── 10. LICENSES & PERMITS ──────────────────────────────────────────────
        (40, lambda d,u,t,a: any(k in u for k in [
             "DE BUSINESS TAX", "CITY OF WILM", "DIV REVE", "LICENSE",
             "PERMIT", "SECRETARY OF STATE"]),
         "Expense: Licenses & Permits", None, "No"),

        # ── 11. INSURANCE ───────────────────────────────────────────────────────
        (40, lambda d,u,t,a: any(k in u for k in [
             "INSUREON", "THE HARTFORD", "TRAVELERS", "HISCOX",
             "NEXT INSURANCE", "LIBERTY MUTUAL", "CHUBB", "MARKEL",
             "INSURANCE", "INSPMTCL"]),
         "Expense: Insurance (General/Liability)", None, "No"),

        # ── 12. SOFTWARE & SUBSCRIPTIONS ────────────────────────────────────────
        (40, lambda d,u,t,a: any(k in u for k in [
             "MSFT", "MICROSOFT", "MSBILL", "ADOBE", "AUTODESK",
             "DROPBOX", "GOOGLE WORKSPACE", "GSUITE", "SLACK",
             "ZOOM", "QUICKBOOKS", "NOTION", "FIGMA", "GITHUB",
             "ANTHROPIC", "OPENAI", "PROTON", "1PASSWORD", "LASTPASS",
             "BLUEBEAM", "ENSCAPE", "LUMION", "RHINO", "FORMIT"]),
         "Expense: Software & Subscriptions", None, "No"),

        # ── 13. PRINTING & REPRO ────────────────────────────────────────────────
        (40, lambda d,u,t,a: any(k in u for k in [
             "FEDEX OFFIC", "FEDEX OFF", "STAPLES", "KINKOS",
             "ALPHAGRAPHICS", "REPROGRAPHICS", "BLUEPRINT", "PLOTTER"]),
         "Expense: Printing & Repro", None, "No"),

        # ── 14. OFFICE SUPPLIES ─────────────────────────────────────────────────
        (45, lambda d,u,t,a: any(k in u for k in [
             "STAPLES", "OFFICE DEPOT", "OFFICEMAX", "AMAZON",
             "TARGET", "COSTCO", "HOME DEPOT", "LOWES"]) and
             "FEDEX" not in u,
         "Expense: Office Supplies", None, "No"),

        # ── 15. HOME DEPOT / LOWES → equipment (small tools) ───────────────────
        (42, lambda d,u,t,a: any(k in u for k in ["HOME DEPOT", "LOWES"]),
         "Expense: Equipment (small tools)", None, "No"),

        # ── 16. ADVERTISING & MARKETING ─────────────────────────────────────────
        (40, lambda d,u,t,a: any(k in u for k in [
             "VISTAPRINT", "MINTED", "CANVA", "LUIGI", "PHOTOGRAPHY",
             "PHOTO", "BRANDING", "WIXSITE", "SQUARESPACE", "GODADDY",
             "NAMECHEAP", "WEBFLOW", "MAILCHIMP", "CONSTANT CONTACT",
             "META ADS", "GOOGLE ADS", "LINKEDIN ADS"]),
         "Expense: Advertising/Marketing", None, ""),

        # ── 17. MEALS ───────────────────────────────────────────────────────────
        (40, lambda d,u,t,a: any(k in u for k in [
             "TST*", "RESTAURANT", "CAFE", "COFFEE", "GRUBHUB",
             "DOORDASH", "UBEREATS", "PACHAMAMA", "CAPERS",
             "STARBUCKS", "PANERA", "CHIPOTLE"]),
         "Expense: Meals (50% typical)", None, "No"),

        # ── 18. TRAVEL ──────────────────────────────────────────────────────────
        (40, lambda d,u,t,a: any(k in u for k in [
             "UBER", "LYFT", "AMTRAK", "UNITED", "DELTA", "SOUTHWEST",
             "AMERICAN AIR", "JETBLUE", "SPIRIT AIR", "FRONTIER",
             "MARRIOTT", "HILTON", "IHG", "HYATT", "AIRBNB",
             "ENTERPRISE", "HERTZ", "AVIS", "BUDGET RENT"]),
         "Expense: Travel", None, "No"),

        # ── 19. VEHICLE / TRANSPORTATION ────────────────────────────────────────
        (40, lambda d,u,t,a: any(k in u for k in [
             "EXXON", "SHELL", "BP ", "WAWA", "SUNOCO", "GULF",
             "PARKWAY", "PARKING", "EZ PASS", "EZPASS", "TURNPIKE",
             "SEPTA", "NJ TRANSIT", "PATCO", "AMTRAK"]),
         "Expense: Vehicle/Transportation", None, "No"),

        # ── 20. PROFESSIONAL FEES (CPA/Legal) ───────────────────────────────────
        (40, lambda d,u,t,a: any(k in u for k in [
             "JETER", "CPA", "ACCOUNTANT", "ATTORNEY", "LAW OFFICE",
             "LEGAL", "NOTARY", "TITLE COMPANY"]),
         "Expense: Professional Fees (CPA/Legal)", None, "No"),

        # ── 21. RENT / WORKSPACE ────────────────────────────────────────────────
        (40, lambda d,u,t,a: any(k in u for k in [
             "RENT", "LEASE", "COWORK", "WEWORK", "REGUS",
             "215 N MARKET", "OFFICE SPACE"]),
         "Expense: Rent/Workspace", None, "No"),

        # ── 22. UTILITIES ───────────────────────────────────────────────────────
        (40, lambda d,u,t,a: any(k in u for k in [
             "DELMARVA", "PECO", "PSEG", "COMCAST", "VERIZON",
             "AT&T", "ELECTRIC", "GAS COMPANY", "WATER BILL"]),
         "Expense: Utilities", None, "No"),

        # ── 23. TRAINING / EDUCATION ────────────────────────────────────────────
        (40, lambda d,u,t,a: any(k in u for k in [
             "UDEMY", "COURSERA", "LINKEDIN LEARN", "AIA ", "CSI ",
             "CONTINUING ED", "SEMINAR", "CONFERENCE", "WORKSHOP"]),
         "Expense: Training/Education", None, "No"),

        # ── 24. BILLPAY to known consultants (1099) ──────────────────────────────
        (50, lambda d,u,t,a: t == "BILLPAY" and any(k in u for k in [
             "PARAGON", "GLOBAL ENGINEERING", "CIVIL ENGINEERING",
             "CARSON STRUCTURAL", "MERESTONE"]),
         "Expense: Consultants (1099)", None, "Yes"),

        # ── 25. BILLPAY to known member names → Distribution ────────────────────
        # (handled separately in member detection pass)
    ]

RULES = make_rules()

# ── PAYEE EXTRACTOR ────────────────────────────────────────────────────────────
BILLPAY_RE = re.compile(r'Online Payment \d+ To (.+?)(?:\s+\d{2}/\d{2})?$', re.I)
CONTRIB_RE  = re.compile(r'Credit Return.*?To (.+?)(?:\s+M4D)?$', re.I)
LEEPRE      = re.compile(r'ORIG CO NAME:([\w\.\s]+)', re.I)

def extract_payee(desc, type_str):
    m = BILLPAY_RE.search(desc)
    if m: return m.group(1).strip()
    m = CONTRIB_RE.search(desc)
    if m: return m.group(1).strip()
    m = LEEPRE.search(desc)
    if m: return m.group(1).strip()
    # Debit card — use first meaningful token(s) before city/state
    if type_str == "DEBIT_CARD":
        clean = re.sub(r'\s+\d{2}/\d{2}$', '', desc).strip()
        clean = re.sub(r'\s{2,}.+$', '', clean).strip()  # strip trailing noise
        return clean[:40]
    return None

def extract_contractor(desc, payee, category):
    """For 1099-flagged rows, return clean contractor name."""
    if not payee: return None
    # Strip M4D suffix, date suffixes
    clean = re.sub(r'\s+M4D\s*\d*/??\d*$', '', payee).strip()
    clean = re.sub(r'\s+\d{2}$', '', clean).strip()
    return clean if clean else None

def is_member(payee):
    if not payee: return None
    for m in MEMBERS:
        if m.lower() in payee.lower():
            return m
    return None

# ── MAIN CATEGORIZE FUNCTION ──────────────────────────────────────────────────
def categorize(row):
    desc   = row['Description'] or ''
    type_  = row['Type'] or ''
    amount = float(row['Amount'])
    upper  = desc.upper()
    debit  = abs(amount) if amount < 0 else 0
    credit = amount      if amount > 0 else 0

    category    = None
    payee       = extract_payee(desc, type_)
    is1099      = ""
    contractor  = None
    flag        = False

    # Member detection pass first
    member = is_member(payee) if type_ in ("BILLPAY", "MISC_CREDIT") else None
    if member:
        if amount < 0:
            category = "Equity: Owner Distribution"
            is1099   = "No"
        elif amount > 0 and type_ == "MISC_CREDIT":
            category = "Equity: Owner Contribution"
            is1099   = "No"
        # If member detected, payee = clean member name
        payee = member

    # Rules pass (only if member didn't already set it)
    if not category:
        for priority, match_fn, cat, payee_hint, is1099_hint in sorted(RULES, key=lambda x: x[0]):
            try:
                if match_fn(row, upper, type_, amount):
                    category = cat
                    if payee_hint and not payee: payee = payee_hint
                    if is1099_hint: is1099 = is1099_hint
                    break
            except Exception:
                continue

    # Contractor name for 1099 rows
    if is1099 == "Yes":
        contractor = extract_contractor(desc, payee, category)

    # Flag anything uncategorized OR that needs human decision
    if not category:
        category = ""
        flag     = True
    elif category in ("Income: Design/Consulting Fees",) and type_ in ("DEPOSIT", "CHECK_DEPOSIT"):
        flag = True  # deposits always need project/client annotation

    return {
        'category':   category,
        'payee':      payee or "",
        'is1099':     is1099,
        'contractor': contractor or "",
        'flag':       flag,
    }

# ── IMPORT INTO WORKBOOK ──────────────────────────────────────────────────────
def import_csv(csv_path, workbook_path, output_path, year_filter=None):
    wb = load_workbook(workbook_path)
    ws = wb['01_Transactions']

    # Find last used row in transactions sheet
    last_row = 2
    for row in ws.iter_rows(min_row=3):
        if any(cell.value is not None for cell in row[:6]):
            last_row = row[0].row

    # Style helpers
    YELLOW  = "FFF2CC"
    ORANGE  = "FCE4D6"
    LGRAY   = "F2F2F2"
    BLK     = "000000"
    DGRAY   = "595959"
    BLUE_IN = "0000FF"
    GREEN_L = "008000"
    RED_W   = "C00000"
    M4D_LT  = "4E9A3A"

    def tb():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s,right=s,top=s,bottom=s)

    def hf(c): return PatternFill("solid", start_color=c, end_color=c)

    # Parse CSV
    with open(csv_path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        csv_rows = list(reader)

    # Deduplicate against existing transactions
    # Build set of (date, amount, first-30-chars-desc) already in sheet
    existing = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] and row[5] is not None:
            try:
                d = row[0]
                if hasattr(d, 'date'): d = d.date()
                existing.add((str(d), str(row[5])[:30]))
            except: pass

    stats = {'imported':0, 'skipped_dup':0, 'skipped_xfer':0, 'flagged':0, 'auto_cat':0}
    import_rows = []

    for csv_row in csv_rows:
        desc   = csv_row['Description'].strip()
        amount = float(csv_row['Amount'])
        type_  = csv_row['Type'].strip()

        # Parse date
        try:
            txn_date = datetime.strptime(csv_row['Posting Date'].strip(), '%m/%d/%Y').date()
        except:
            continue

        # Year filter
        if year_filter and txn_date.year != year_filter:
            continue

        # Dedup check
        dup_key = (str(txn_date), str(amount)[:30])
        if dup_key in existing:
            stats['skipped_dup'] += 1
            continue

        result = categorize(csv_row)

        debit  = abs(amount) if amount < 0 else 0
        credit = amount      if amount > 0 else 0
        q_num  = (txn_date.month - 1) // 3 + 1

        import_rows.append({
            'date':       txn_date,
            'account':    ACCOUNT_NAME,
            'banktype':   type_,
            'desc':       desc[:120],
            'payee':      result['payee'][:60] if result['payee'] else "",
            'amount':     amount,
            'debit':      debit,
            'credit':     credit,
            'category':   result['category'],
            'taxline':    "",
            'is1099':     result['is1099'],
            'contractor': result['contractor'],
            'project':    "38 E 23rd / Claymont / Harrison" if "L.E.E.P" in desc else "",
            'memo':       "",
            'receipt':    "",
            'reviewed':   "",
            'quarter':    f"Q{q_num}",
            'month':      txn_date.month,
            'flag':       result['flag'],
            'category_auto': bool(result['category']),
        })

        existing.add(dup_key)
        if result['flag']: stats['flagged'] += 1
        if result['category'] and not result['flag']: stats['auto_cat'] += 1
        if result['category'] == "Transfer: Between Accounts": stats['skipped_xfer'] += 1
        stats['imported'] += 1

    # Sort by date
    import_rows.sort(key=lambda x: x['date'])

    # Write rows to workbook
    for ir in import_rows:
        r = last_row + 1
        last_row = r

        bg = ORANGE if ir['flag'] else LGRAY

        vals = [
            (1,  ir['date'],       'MM/DD/YYYY',                    BLK),
            (2,  ir['account'],    '@',                             DGRAY),
            (3,  ir['banktype'],   '@',                             DGRAY),
            (4,  ir['desc'],       '@',                             DGRAY),
            (5,  ir['payee'],      '@',                             BLK),
            (6,  ir['amount'],     '$#,##0.00;($#,##0.00);"-"',    BLK if ir['amount']>0 else RED_W),
            (7,  ir['debit'],      '$#,##0.00;($#,##0.00);"-"',    BLK),
            (8,  ir['credit'],     '$#,##0.00;($#,##0.00);"-"',    BLK),
            (9,  ir['category'],   '@',                             BLK if ir['category'] else BLUE_IN),
            (10, ir['taxline'],    '@',                             DGRAY),
            (11, ir['is1099'],     '@',                             BLK),
            (12, ir['contractor'], '@',                             BLK),
            (13, ir['project'],    '@',                             BLK),
            (14, ir['memo'],       '@',                             BLK),
            (15, ir['receipt'],    '@',                             BLK),
            (16, ir['reviewed'],   '@',                             BLK),
            (17, f'="Q"&INT((MONTH(A{r})+2)/3)', '@',              DGRAY),
            (18, f'=MONTH(A{r})',  '0',                             DGRAY),
        ]

        for col, val, fmt, color in vals:
            cell = ws.cell(r, col, value=val)
            cell.font = Font(name="Arial", size=10, color=color)
            cell.number_format = fmt
            cell.border = tb()
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(col==4))
            # Flag rows: orange background on category + payee columns
            if ir['flag'] and col in [5, 9]:
                cell.fill = hf(ORANGE)
            elif col in [9] and ir['category_auto']:
                cell.fill = hf("E2EFDA")  # light green = auto-categorized
            elif col in [9] and not ir['category']:
                cell.fill = hf(ORANGE)    # orange = needs manual entry

        ws.row_dimensions[r].height = 16

    # Add a separator/header row before imported block so it's easy to find
    # (Insert a note in col D of last_row-len+1 if we wrote anything)
    if import_rows:
        sep_row = last_row - len(import_rows) + 1 - 1
        if sep_row >= 3:
            note = ws.cell(sep_row, 4)
            # Only add if that row is empty
            if not note.value:
                pass  # don't stomp existing data

    wb.save(output_path)
    return stats, import_rows

# ── RUN ───────────────────────────────────────────────────────────────────────
def build_output_path(workbook_path: str) -> str:
    """Default output path: same dir as workbook, with _imported suffix."""
    base, ext = os.path.splitext(workbook_path)
    return f"{base}_imported{ext}"


def parse_args():
    parser = argparse.ArgumentParser(
        description="M4D Chase CSV Importer — maps Chase bank exports into the M4D bookkeeping workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 M4D_Chase_Importer.py Chase6378_Activity_20260130.CSV --workbook Merge4Design_2026_Bookkeeping_v1.xlsx
  python3 M4D_Chase_Importer.py Chase6378_Activity_20260130.CSV --workbook ./bookkeeping/Merge4Design_2026_Bookkeeping_v1.xlsx --year 2026
  python3 M4D_Chase_Importer.py Chase6378_Activity_20260130.CSV --workbook workbook.xlsx --output reviewed/workbook_jan.xlsx
        """,
    )
    parser.add_argument(
        "csv",
        metavar="CSV_FILE",
        help="Chase CSV export file (downloaded from chase.com)",
    )
    parser.add_argument(
        "--workbook", "-w",
        required=True,
        metavar="XLSX_FILE",
        help="Path to the M4D bookkeeping workbook (.xlsx)",
    )
    parser.add_argument(
        "--output", "-o",
        metavar="OUTPUT_FILE",
        default=None,
        help="Output path for the imported workbook (default: <workbook>_imported.xlsx)",
    )
    parser.add_argument(
        "--year", "-y",
        type=int,
        default=None,
        metavar="YYYY",
        help="Filter to a single tax year (e.g. 2026). Omit to import all years in the file.",
    )
    parser.add_argument(
        "--account",
        default=ACCOUNT_NAME,
        metavar="ACCT",
        help=f"Account label written into the Account column (default: {ACCOUNT_NAME!r})",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()

    csv_path      = args.csv
    workbook_path = args.workbook
    output_path   = args.output or build_output_path(workbook_path)
    year_filter   = args.year

    # Validate inputs
    if not os.path.isfile(csv_path):
        print(f"ERROR: CSV not found: {csv_path}", file=sys.stderr)
        sys.exit(1)
    if not os.path.isfile(workbook_path):
        print(f"ERROR: Workbook not found: {workbook_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Chase CSV:    {csv_path}")
    print(f"Workbook:     {workbook_path}")
    print(f"Output:       {output_path}")
    print(f"Year filter:  {year_filter or 'ALL'}")
    print()

    stats, rows = import_csv(csv_path, workbook_path, output_path, year_filter)

    print(f"Results:")
    print(f"  Total CSV rows:        {stats['imported'] + stats['skipped_dup']}")
    print(f"  Imported:              {stats['imported']}")
    print(f"  Skipped (duplicates):  {stats['skipped_dup']}")
    print(f"  Auto-categorized:      {stats['auto_cat']}")
    print(f"  Flagged for review:    {stats['flagged']}")
    print(f"  Transfers (auto):      {stats['skipped_xfer']}")
    print()

    # Print review list
    needs_review = [r for r in rows if r['flag']]
    if needs_review:
        print(f"── NEEDS MANUAL REVIEW ({len(needs_review)} rows) ──────────────────")
        for r in needs_review:
            cat_status = f"CAT: {r['category']!r}" if r['category'] else "NO CATEGORY"
            print(f"  {r['date']}  {r['amount']:>10.2f}  {r['desc'][:55]:<55}  {cat_status}")
    else:
        print("All rows auto-categorized. ✓")

    # Print category summary
    from collections import Counter
    cat_counts = Counter(r['category'] for r in rows if r['category'])
    print()
    print("── CATEGORY BREAKDOWN ──────────────────────────────────")
    for cat, count in sorted(cat_counts.items()):
        total = sum(abs(r['amount']) for r in rows if r['category']==cat)
        print(f"  {count:>3}x  {cat:<45}  ${total:>10,.2f}")

    print(f"\nOutput: {output_path}")
